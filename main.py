import requests
import json
from openpyxl import load_workbook
import socketserver as socketserver
from http.server import SimpleHTTPRequestHandler as RequestHandler
import datetime
import line_bot_lib

# get auth_token
workbook = load_workbook('data.xlsx')
sheetAuthToken = workbook['Sheet']
auth_token = sheetAuthToken.cell(row=2, column=1).value

class MyHandler(RequestHandler):
    def do_POST(self):
        varLen = int(self.headers['Content-Length'])
        if varLen > 0:
            post_data = self.rfile.read(varLen)
            data = json.loads(post_data)
            userID = data['events'][0]['source']['userId']
            replyToken = data['events'][0]['replyToken']  # need this token to reply
            inputText = data['events'][0]['message']['text']  # sent content

        reply = line_bot_lib.get_reply(inputText)

        message = {'replyToken': replyToken, 'messages': reply}
        # 將資料存到log.xlsx
        inputLog = [datetime.datetime.now().strftime('%Y/%m/%d %H:%M:%S'), userID, inputText, str(reply)]
        logWorkbook = load_workbook('log.xlsx')
        logSheet = logWorkbook['log']
        row = 1
        while True:  # 找輸入的列
            if logSheet.cell(row=row, column=1).value == None:
                break
            row += 1
        for i in range(4):  # 把資料填入
            logSheet.cell(row=row, column=i + 1).value = inputLog[i]
        logWorkbook.save('log.xlsx')  # 儲存檔案

        hed = {'Authorization': 'Bearer ' + auth_token}
        url = 'https://api.line.me/v2/bot/message/reply'
        self.send_response(200)
        self.end_headers()
        requests.post(url, json=message, headers=hed)


socketserver.TCPServer.allow_reuse_address = True
httpd = socketserver.TCPServer(('0.0.0.0', 8888), MyHandler)
try:
    httpd.serve_forever()
except:
    print('Closing the server.')
    httpd.server_close()
